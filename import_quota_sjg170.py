#!/usr/bin/env python3
"""
导入土方与地基基础工程消耗量标准（SJG170-2024）。
Excel 为结构化导出格式，含独立的标准信息/定额子目/费用构成/工料机消耗量四个 Sheet。

用法：
  python import_quota_sjg170.py <excel_path>
  python import_quota_sjg170.py <excel_path> --force
"""

import argparse
import os
import sys
import datetime

import openpyxl
from dotenv import load_dotenv

from db.connection import get_connection
from importer.quota_loader import init_schema

load_dotenv()

RESOURCE_TYPE_MAP = {
    'labor': '人工',
    'material': '材料',
    'machine': '机械',
    'other': '其他',
}


def _clean(v):
    """清理单元格值：None 保持 None，字符串去首尾空白。"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_workbook(path: str) -> dict:
    """解析结构化 Excel，返回标准信息、章节、子目、工料机数据。"""
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── 标准信息 ──────────────────────────────────────────
    info = {}
    ws_info = wb['标准信息']
    for row in ws_info.iter_rows(min_row=2, values_only=True):
        key, val = row[0], row[1]
        if key:
            info[str(key).strip()] = val

    standard_code = str(info.get('标准编码', 'SJG170-2024')).strip()
    standard_name = str(info.get('标准名称', '土石方与地基基础工程消耗量标准')).strip()
    region = str(info.get('地区', '深圳')).strip()

    # ── 定额子目 ──────────────────────────────────────────
    ws_items = wb['定额子目']
    # 列：编码, 名称, 规格, 单位, 章节编码, 章节名称, 工作内容, 来源起始行, 来源结束行
    items_raw = []
    chapters = {}
    for row in ws_items.iter_rows(min_row=2, values_only=True):
        code = _clean(row[0])
        if not code:
            continue
        item_name = _clean(row[1])
        variant_desc = _clean(row[2])
        unit = _clean(row[3])
        chapter_code = _clean(row[4])
        chapter_name = _clean(row[5])
        work_content = _clean(row[6])

        if chapter_code and chapter_name:
            chapters[chapter_code] = chapter_name

        items_raw.append({
            'item_code': code,
            'item_name': item_name,
            'variant_desc': variant_desc,
            'unit': unit,
            'chapter_code': chapter_code,
            'work_content': work_content,
        })

    # ── 费用构成 ──────────────────────────────────────────
    ws_cost = wb['费用构成']
    # 列：编码, 名称, 规格, 单位, 全费用综合单价, 参考综合单价, 人工费, 材料费, 机械费,
    #      管理费, 利润, 安全文明施工措施费, 规费, 税金, 价格期
    costs = {}
    for row in ws_cost.iter_rows(min_row=2, values_only=True):
        code = _clean(row[0])
        if not code:
            continue
        costs[code] = {
            'total_unit_price': _to_float(row[4]),
            'unit_price': _to_float(row[5]),
            'labor_cost': _to_float(row[6]),
            'material_cost': _to_float(row[7]),
            'machine_cost': _to_float(row[8]),
            'management_fee': _to_float(row[9]),
            'profit': _to_float(row[10]),
            'safety_fee': _to_float(row[11]),
            'statutory_fee': _to_float(row[12]),
            'tax': _to_float(row[13]),
        }

    # ── 工料机消耗量 ──────────────────────────────────────
    ws_res = wb['工料机消耗量']
    # 列：子目编码, 子目名称, 子目规格, 资源类型, 资源名称, 单位, 消耗量, 参考价, 来源行
    resources = {}
    for row in ws_res.iter_rows(min_row=2, values_only=True):
        code = _clean(row[0])
        if not code:
            continue
        res_type_en = _clean(row[3]) or 'other'
        res_type = RESOURCE_TYPE_MAP.get(res_type_en, res_type_en)
        res_name = _clean(row[4])
        res_unit = _clean(row[5])
        qty = _to_float(row[6])
        ref_price = _to_float(row[7])

        if code not in resources:
            resources[code] = []
        resources[code].append({
            'resource_type': res_type,
            'resource_name': res_name or res_type,
            'unit': res_unit,
            'quantity': qty,
            'ref_price': ref_price,
        })

    return {
        'standard_code': standard_code,
        'standard_name': standard_name,
        'region': region,
        'chapters': chapters,
        'items_raw': items_raw,
        'costs': costs,
        'resources': resources,
    }


def import_to_db(conn, data: dict, source_file: str, force: bool = False):
    # 确保 schema 已初始化
    init_schema(conn)

    standard_code = data['standard_code']

    # 检查是否已存在
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM quota_standards WHERE standard_code = %s", (standard_code,))
        existing = cur.fetchone()

    if existing:
        if not force:
            print(f'[跳过] 标准 {standard_code} 已存在（id={existing[0]}）。使用 --force 强制重新导入。')
            return
        print(f'[覆盖] 删除旧数据 (id={existing[0]}) ...')
        with conn.cursor() as cur:
            cur.execute("DELETE FROM quota_standards WHERE id = %s", (existing[0],))
        conn.commit()

    # 检查 quota_standards 是否有 region 列，没有就补加
    with conn.cursor() as cur:
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name='quota_standards' AND column_name='region'
        """)
        if not cur.fetchone():
            cur.execute("ALTER TABLE quota_standards ADD COLUMN region VARCHAR(100)")
            conn.commit()
            print('  已追加 quota_standards.region 列')

    # 插入标准记录
    base_date = datetime.date(2023, 8, 1)
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO quota_standards
                (standard_code, name, base_date, source_file, region)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id
        """, (standard_code, data['standard_name'], base_date, source_file, data['region']))
        standard_id = cur.fetchone()[0]
    conn.commit()
    print(f'  标准记录 id={standard_id}')

    # 插入章节
    chapter_id_map = {}
    with conn.cursor() as cur:
        for i, (code, name) in enumerate(sorted(data['chapters'].items())):
            cur.execute("""
                INSERT INTO quota_chapters
                    (standard_id, code, name, level, parent_id, sort_order)
                VALUES (%s, %s, %s, 1, NULL, %s)
                RETURNING id
            """, (standard_id, code, name, i + 1))
            chapter_id_map[code] = cur.fetchone()[0]
    conn.commit()
    print(f'  写入 {len(chapter_id_map)} 个章节')

    # 插入子目 + 工料机
    n_items = 0
    n_resources = 0
    with conn.cursor() as cur:
        for item in data['items_raw']:
            code = item['item_code']
            cost = data['costs'].get(code, {})
            chapter_id = chapter_id_map.get(item['chapter_code'])

            cur.execute("""
                INSERT INTO quota_items
                    (standard_id, chapter_id, item_code, item_name, variant_desc,
                     unit, work_content,
                     total_unit_price, unit_price,
                     labor_cost, material_cost, machine_cost,
                     management_fee, profit, safety_fee, statutory_fee, tax)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                standard_id, chapter_id, code,
                item['item_name'], item['variant_desc'],
                item['unit'], item['work_content'],
                cost.get('total_unit_price'), cost.get('unit_price'),
                cost.get('labor_cost'), cost.get('material_cost'), cost.get('machine_cost'),
                cost.get('management_fee'), cost.get('profit'),
                cost.get('safety_fee'), cost.get('statutory_fee'), cost.get('tax'),
            ))
            item_id = cur.fetchone()[0]
            n_items += 1

            for res in data['resources'].get(code, []):
                cur.execute("""
                    INSERT INTO quota_resources
                        (item_id, resource_type, resource_name, unit, quantity, ref_price)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (item_id, res['resource_type'], res['resource_name'],
                      res['unit'], res['quantity'], res['ref_price']))
                n_resources += 1

    conn.commit()
    print(f'  写入 {n_items} 个子目，{n_resources} 条工料机')


def main():
    parser = argparse.ArgumentParser(description='导入土方与地基基础工程消耗量标准（结构化Excel）')
    parser.add_argument('excel_path', help='Excel 文件路径')
    parser.add_argument('--force', action='store_true', help='若已存在则强制重新导入')
    args = parser.parse_args()

    if not os.path.isfile(args.excel_path):
        print(f'[错误] 文件不存在: {args.excel_path}', file=sys.stderr)
        sys.exit(1)

    print(f'解析 {args.excel_path} ...')
    data = parse_workbook(args.excel_path)
    print(f'  {data["standard_name"]}（{data["standard_code"]}）')
    print(f'  {len(data["chapters"])} 个章节，{len(data["items_raw"])} 个子目')

    source_file = os.path.basename(args.excel_path)
    conn = get_connection()
    try:
        print('写入数据库 ...')
        import_to_db(conn, data, source_file, args.force)
        print('导入完成。')
    except Exception as e:
        conn.rollback()
        print(f'[错误] {e}', file=sys.stderr)
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
