import re
import os
from dataclasses import dataclass
from typing import List, Optional, Tuple
from openpyxl import load_workbook


@dataclass
class SheetMeta:
    sheet_index: int
    sheet_name: str
    category_group: str


@dataclass
class PriceItem:
    sheet_index: int
    sequence_no: Optional[int]
    material_code: Optional[str]
    material_name: str
    specification: Optional[str]
    unit: Optional[str]
    price_yuan: Optional[float]
    coefficient: Optional[float]
    calculation_formula: Optional[str]
    remarks: Optional[str]


def parse_filename(path: str) -> Tuple[int, int, int]:
    name = os.path.basename(path)
    m = re.search(r'(\d{4})-(\d{2})-(\d+)深圳信息价', name)
    if not m:
        raise ValueError(
            f"文件名格式不符合预期: {name}\n"
            "期望格式: YYYY-MM-版本深圳信息价.xlsx，例如: 2026-04-0深圳信息价.xlsx"
        )
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def _str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _decimal(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def parse_workbook(path: str) -> Tuple[List[SheetMeta], List[PriceItem]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets_meta: List[SheetMeta] = []
    items: List[PriceItem] = []

    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            continue

        # 从 sheet 标签名提取大类，格式：`建筑材料-1.黑色及有色金属` → `建筑材料`
        category_group = sheet_name.split("-")[0].strip() if "-" in sheet_name else sheet_name

        sheets_meta.append(SheetMeta(
            sheet_index=idx,
            sheet_name=sheet_name,
            category_group=category_group,
        ))

        # Row 1 (Excel row 2): column headers — skip
        # Row 2+ (Excel row 3+): data
        for row in rows[2:]:
            # Columns A-I: 序号 材料编码 材料名称 型号规格 单位 价格 系数 计算公式 备注
            if len(row) < 3:
                continue
            material_name = _str(row[2])
            if not material_name:
                continue

            items.append(PriceItem(
                sheet_index=idx,
                sequence_no=_int(row[0] if len(row) > 0 else None),
                material_code=_str(row[1] if len(row) > 1 else None),
                material_name=material_name,
                specification=_str(row[3] if len(row) > 3 else None),
                unit=_str(row[4] if len(row) > 4 else None),
                price_yuan=_decimal(row[5] if len(row) > 5 else None),
                coefficient=_decimal(row[6] if len(row) > 6 else None),
                calculation_formula=_str(row[7] if len(row) > 7 else None),
                remarks=_str(row[8] if len(row) > 8 else None),
            ))

    wb.close()
    return sheets_meta, items
