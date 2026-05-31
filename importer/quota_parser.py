"""
解析深圳市建筑消耗量标准 Excel 文件。

Excel 结构：单 Sheet，前 ~246 行为封面/前言/目录，从 Row 247 开始为子目数据块。
每个子目块包含：章节标题、工作内容、子目编号/名称/变体、价格构成、工料机消耗量。
"""

import re
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Tuple
from openpyxl import load_workbook


# ---------- 数据模型 ----------

@dataclass
class ChapterInfo:
    code: str
    name: str
    level: int
    sort_order: int


@dataclass
class ResourceItem:
    resource_type: str      # '人工' | '材料' | '机械'
    resource_name: str
    unit: Optional[str]
    quantity: Optional[float]
    ref_price: Optional[float]


@dataclass
class QuotaItem:
    item_code: str
    item_name: str
    variant_desc: Optional[str]
    unit: Optional[str]
    work_content: Optional[str]
    chapter_code: Optional[str]
    total_unit_price: Optional[float]
    unit_price: Optional[float]
    labor_cost: Optional[float]
    material_cost: Optional[float]
    machine_cost: Optional[float]
    management_fee: Optional[float]
    profit: Optional[float]
    safety_fee: Optional[float]
    statutory_fee: Optional[float]
    tax: Optional[float]
    source_row: int
    resources: List[ResourceItem] = field(default_factory=list)


# ---------- 辅助函数 ----------

def _str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _decimal(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r'[\s,，]', '', str(val))
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _clean_code(val) -> Optional[str]:
    """清除子目编号中的空格，如 '010001 - 7' → '010001-7'"""
    if val is None:
        return None
    s = re.sub(r'\s+', '', str(val))
    return s if s else None


def _extract_unit(text: str) -> Optional[str]:
    """从工作内容行末尾提取计量单位，如 '...  单位 :   10m3' → '10m3'"""
    if not text:
        return None
    m = re.search(r'单\s*位\s*[：:]\s*(.+?)(?:\s*$)', text.strip())
    if m:
        return re.sub(r'\s+', '', m.group(1)).strip()
    return None


def _is_quota_code(val) -> bool:
    """判断一个值是否像子目编号字符串（如 010001-7）"""
    if val is None:
        return False
    s = re.sub(r'\s+', '', str(val))
    return bool(re.match(r'^\d{6}-\d+$', s))


def _is_quota_suffix_int(val) -> bool:
    """判断是否是被 Excel 解析为负整数的编号后缀（如 -2 代表 xxx-2）"""
    return isinstance(val, int) and val < 0


def _parse_chapter_line(text: str) -> Tuple[Optional[str], Optional[str], int]:
    """
    解析章节标题文本，返回 (code, name, level)。
    例：'1 .3 . 1    砖    砌    体' → ('1.3.1', '砖砌体', 3)
         '1    砌筑工程' → ('1', '砌筑工程', 1)
    """
    text = text.strip()
    compact = re.sub(r'\s+', '', text)
    m = re.match(r'^([\d.]+)(.+)$', compact)
    if m:
        raw_code = m.group(1).rstrip('.')
        name = m.group(2).strip()
        level = raw_code.count('.') + 1
        return raw_code, name, level
    return None, text, 1


# 可能出现子目编号的列（0-indexed）
CANDIDATE_VARIANT_COLS = [13, 14, 17, 18, 19, 21, 22, 23]


def _detect_variant_cols(row: tuple) -> List[int]:
    """从子目编号行检测变体列（0-indexed），包括字符串编号和负整数后缀"""
    return [
        c for c in CANDIDATE_VARIANT_COLS
        if c < len(row) and (
            _is_quota_code(row[c]) or _is_quota_suffix_int(row[c])
        )
    ]


# ---------- 状态机解析器 ----------

class _Parser:
    def __init__(self):
        self.chapters: List[ChapterInfo] = []
        self.items: List[QuotaItem] = []
        self._chapter_sort = 0
        self._seen_chapters = set()

        # 当前全局状态
        self._current_chapter_code: Optional[str] = None
        self._work_content: Optional[str] = None
        self._unit: Optional[str] = None
        self._last_prefix: Optional[str] = None  # 最近一次有效的6位编号前缀（如 "010001"）

        # 当前子目块状态
        self._variant_cols: List[int] = []
        self._item_name: Optional[str] = None
        self._variant_descs: Dict[int, List[str]] = {}
        self._prices: Dict[int, Dict] = {}
        self._resources: Dict[int, List[ResourceItem]] = {}
        self._res_type: str = '材料'
        self._in_price = False
        self._in_resource = False
        self._block_row = 0

    def _flush(self):
        if not self._variant_cols or not self._item_name:
            return
        for col in self._variant_cols:
            raw_code = self._prices.get(col, {}).get('__code__')
            # 重建编号：字符串直接清洗，负整数用已知前缀重建
            if _is_quota_suffix_int(raw_code):
                if self._last_prefix:
                    code = f"{self._last_prefix}-{abs(raw_code)}"
                else:
                    continue  # 无法推断前缀，跳过
            else:
                code = _clean_code(raw_code)
                if not code:
                    continue
            desc_parts = self._variant_descs.get(col, [])
            vdesc = ' / '.join(p for p in desc_parts if p) or None
            p = self._prices.get(col, {})
            self.items.append(QuotaItem(
                item_code=code,
                item_name=self._item_name,
                variant_desc=vdesc,
                unit=self._unit,
                work_content=self._work_content,
                chapter_code=self._current_chapter_code,
                total_unit_price=p.get('total'),
                unit_price=p.get('unit'),
                labor_cost=p.get('labor'),
                material_cost=p.get('material'),
                machine_cost=p.get('machine'),
                management_fee=p.get('management'),
                profit=p.get('profit'),
                safety_fee=p.get('safety'),
                statutory_fee=p.get('statutory'),
                tax=p.get('tax'),
                source_row=self._block_row,
                resources=list(self._resources.get(col, [])),
            ))

    def _reset_block(self):
        self._variant_cols = []
        self._item_name = None
        self._variant_descs = {}
        self._prices = {}
        self._resources = {}
        self._res_type = '材料'
        self._in_price = False
        self._in_resource = False
        self._block_row = 0

    def _add_chapter(self, code, name, level):
        key = (code or '') + name
        if key in self._seen_chapters:
            return
        self._seen_chapters.add(key)
        self._chapter_sort += 1
        self.chapters.append(ChapterInfo(
            code=code or '',
            name=name,
            level=level,
            sort_order=self._chapter_sort,
        ))
        if level <= 2 and code:
            self._current_chapter_code = code

    def feed(self, row_idx: int, raw_row: tuple):
        row = raw_row + (None,) * max(0, 30 - len(raw_row))

        c1  = _str(row[0])
        c3  = _str(row[2])
        c4  = _str(row[3])
        c12 = _str(row[11])
        c27 = _decimal(row[26])

        # ── 章节标题行：col12 有内容，col1 / col14 / col15 为空 ──
        if c12 and not c1 and row[13] is None and row[14] is None:
            # 排除：单位词短串、纯数字、工料机参考价格表格行（col27 有数值）
            c12_compact = re.sub(r'\s+', '', c12)
            # 必须包含至少2个汉字，且长度足够
            has_cn = len(re.findall(r'[一-鿿]', c12_compact)) >= 2
            has_ref_price = row[26] is not None  # col27 有数值说明是工料机行
            if has_cn and not has_ref_price:
                code, name, level = _parse_chapter_line(c12)
                if name and len(re.findall(r'[一-鿿]', name)) >= 2:
                    self._add_chapter(code, name, level)
            return

        # col10 小节标题（如 '2    建筑物非泵送现浇混凝土'）
        c10 = _str(row[9])
        if c10 and not c1 and not c12 and row[14] is None and row[13] is None:
            c10_compact = re.sub(r'\s+', '', c10)
            has_cn = len(re.findall(r'[一-鿿]', c10_compact)) >= 3
            if has_cn:
                code, name, level = _parse_chapter_line(c10)
                if name:
                    self._add_chapter(code, name, max(level, 3))
            return

        # ── 工作内容行 ──
        if c1 and '工作内容' in c1:
            self._flush()
            self._reset_block()
            self._work_content = c1
            self._unit = _extract_unit(c1)
            return

        # ── 子目编号行 ──
        if c1 and re.search(r'子\s*目\s*编\s*号', c1):
            vcols = _detect_variant_cols(row)
            if vcols:
                # 若当前有未 flush 的块，先 flush（两个子目块共享工作内容时无中间 工作内容 行）
                if self._variant_cols and self._item_name:
                    self._flush()
                self._variant_cols = vcols
                self._block_row = row_idx
                self._prices = {c: {'__code__': row[c]} for c in vcols}
                self._variant_descs = {c: [] for c in vcols}
                self._resources = {c: [] for c in vcols}
                self._in_price = False
                self._in_resource = False
                self._item_name = None
                # 从本行字符串编号更新前缀（如 '010001-7' → '010001'）
                for c in vcols:
                    if _is_quota_code(row[c]):
                        m = re.match(r'^(\d{6})', re.sub(r'\s+', '', str(row[c])))
                        if m:
                            self._last_prefix = m.group(1)
                            break
            return

        if not self._variant_cols:
            return

        # ── 子目名称行 ──
        if c1 and re.search(r'子\s*目\s*名\s*称', c1):
            for c in self._variant_cols:
                v = _str(row[c])
                if v:
                    self._item_name = v
                    break
            return

        # ── 变体描述行（价格前、col1 为空）──
        if not self._in_price and not self._in_resource and not c1:
            for c in self._variant_cols:
                v = _str(row[c])
                if v:
                    self._variant_descs[c].append(v)
            return

        # ── 全费用综合单价行 ──
        if c1 and '全费用' in c1 and '构成' not in c1:
            self._in_price = True
            for c in self._variant_cols:
                self._prices[c]['total'] = _decimal(row[c])
            return

        # ── 价格构成段 ──
        if self._in_price and not self._in_resource:
            label4 = c4 or ''
            label3 = c3 or ''
            if c1 and '综合单价构成' in c1:
                for c in self._variant_cols:
                    self._prices[c]['unit'] = _decimal(row[c])
                return
            if '人工费' in label4 and c4:
                for c in self._variant_cols:
                    self._prices[c]['labor'] = _decimal(row[c])
                return
            if '材料费' in label4 and c4:
                for c in self._variant_cols:
                    self._prices[c]['material'] = _decimal(row[c])
                return
            if '机械费' in label4 and c4:
                for c in self._variant_cols:
                    self._prices[c]['machine'] = _decimal(row[c])
                return
            if '管理费' in label4 and c4:
                for c in self._variant_cols:
                    self._prices[c]['management'] = _decimal(row[c])
                return
            if '利润' in label4 and c4:
                for c in self._variant_cols:
                    self._prices[c]['profit'] = _decimal(row[c])
                return
            if '安全' in label3 and c3:
                for c in self._variant_cols:
                    self._prices[c]['safety'] = _decimal(row[c])
                return
            if '规费' in label3 and c3:
                for c in self._variant_cols:
                    self._prices[c]['statutory'] = _decimal(row[c])
                return
            if '税金' in label3 and c3:
                for c in self._variant_cols:
                    self._prices[c]['tax'] = _decimal(row[c])
                return

        # ── 工料机段开始 ──
        if c1 and '工料机名称' in c1:
            self._in_resource = True
            self._res_type = '材料'
            return

        if not self._in_resource:
            return

        # ── 工料机段：确定资源类型 ──
        if c1:
            if re.match(r'^人\s*工', c1):
                self._res_type = '人工'
            elif re.match(r'^材\s*[料 　]', c1):
                self._res_type = '材料'
            elif re.match(r'^机\s*[械 　]', c1):
                self._res_type = '机械'

        # 资源名在 col3
        res_name = c3
        if not res_name:
            return
        res_unit = _str(row[11])

        for c in self._variant_cols:
            qty = _decimal(row[c])
            if qty is None:
                continue
            self._resources[c].append(ResourceItem(
                resource_type=self._res_type,
                resource_name=res_name,
                unit=res_unit,
                quantity=qty,
                ref_price=c27,
            ))

    def finish(self):
        self._flush()


# ---------- 公开接口 ----------

def parse_quota_workbook(path: str) -> Tuple[List[ChapterInfo], List[QuotaItem]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    parser = _Parser()
    for row_idx, raw_row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx < 247:
            continue
        parser.feed(row_idx, raw_row)
    wb.close()
    parser.finish()
    return parser.chapters, parser.items
