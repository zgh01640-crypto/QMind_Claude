"""
解析 E.1 分部分项工程项目清单计价表 Excel。

返回:
  project_info: dict  {project_name, bid_section}
  sections: list[dict]  [{seq, section_name}]
  items: list[dict]  [{item_seq, item_code, item_name, item_description,
                        unit, quantity, unit_price, total_price, provisional_price,
                        section_seq}]  # section_seq 与 sections 的 seq 对应
"""

import re
import openpyxl


_SKIP_PATTERNS = re.compile(r'^(本页小计|合计|分部小计)$')


def _parse_project_info(ws):
    """从第1、2行提取工程名和标段。"""
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    project_name = ''
    bid_section = ''
    for cell_val in row2:
        if cell_val is None:
            continue
        s = str(cell_val).strip()
        if s.startswith('工程名称'):
            project_name = re.sub(r'^工程名称[：:]\s*', '', s)
        elif s.startswith('标段'):
            bid_section = re.sub(r'^标段[：:]\s*', '', s)
    return {'project_name': project_name, 'bid_section': bid_section}


def _is_section_row(row):
    """分部行：序号为空，第3列有内容，第2/4/5/6列均为空。"""
    seq, code, name, desc, unit, qty = row[0], row[1], row[2], row[3], row[4], row[5]
    if seq is not None and str(seq).strip():
        return False
    if not name or not str(name).strip():
        return False
    if code or desc or unit or (qty is not None and str(qty).strip()):
        return False
    return True


def _is_item_row(row):
    """清单行：序号为整数或可转为整数的字符串，有项目编码。"""
    seq, code = row[0], row[1]
    if seq is None or code is None:
        return False
    try:
        int(str(seq).strip())
    except (ValueError, AttributeError):
        return False
    return bool(str(code).strip())


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def parse_boq_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    project_info = _parse_project_info(ws)

    sections = []
    items = []
    current_section_seq = 0
    section_seq_counter = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        # 取前9列，不足则补 None
        r = list(row) + [None] * 9
        r = r[:9]
        seq, code, name, desc, unit, qty, unit_price, total_price, prov_price = r

        name_s = str(name).strip() if name else ''

        # 跳过小计/合计
        if name_s and _SKIP_PATTERNS.match(name_s):
            continue
        # 跳过序号列是"本页小计"/"合计"
        if seq and _SKIP_PATTERNS.match(str(seq).strip()):
            continue

        row_data = [seq, code, name, desc, unit, qty]

        if _is_section_row(row_data):
            section_seq_counter += 1
            sections.append({'seq': section_seq_counter, 'section_name': name_s})
            current_section_seq = section_seq_counter
        elif _is_item_row(row_data):
            items.append({
                'item_seq': int(str(seq).strip()),
                'item_code': str(code).strip(),
                'item_name': str(name).strip(),
                'item_description': str(desc).strip() if desc else None,
                'unit': str(unit).strip() if unit else None,
                'quantity': _to_float(qty),
                'unit_price': _to_float(unit_price),
                'total_price': _to_float(total_price),
                'provisional_price': _to_float(prov_price),
                'section_seq': current_section_seq,
            })

    return project_info, sections, items
