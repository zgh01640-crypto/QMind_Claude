#!/usr/bin/env python3
"""
导入人工套定额的工程量清单（含子定额）Excel。

Excel格式：单Sheet，8列：序号/项目代码/项目名称/项目规格/单位/数量/综合单价/合计
行类型：分部标题 | BOQ清单项（C1为整数）| 定额子目行（C1空+C2有编码）| 汇总行（跳过）

用法：
  python import_manual_boq.py <excel_path> [--tag TAG] [--force]
"""

import argparse
import os
import re
import sys

import openpyxl
from dotenv import load_dotenv

from db.connection import get_connection

load_dotenv()

_QUOTA_CODE_RE = re.compile(r'^\d{6}[-\d+*.\s]+')  # 简单码: 120001-11；公式: 120001-214+...
_SIMPLE_CODE_RE = re.compile(r'^\d{6}-\d+$')        # 纯简单码，可查 quota_items


def _clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _is_skip_row(c3: str | None) -> bool:
    if not c3:
        return False
    keywords = ['小计', '合计', '本页', '分部分项工程', '其中']
    return any(kw in c3 for kw in keywords)


def parse_workbook(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 提取工程名称（行2）
    row2 = [_clean(ws.cell(2, c).value) for c in range(1, 9)]
    project_name = row2[2] or row2[0] or '未命名工程'
    bid_section = None
    if row2[3]:
        bid_section = row2[3]

    sections = []
    items = []
    quotas = []  # list of (item_index, quota_dict)

    current_section_seq = 0
    current_section_name = None
    current_item_index = None  # index into items[]

    for row in ws.iter_rows(min_row=4, values_only=True):
        c1, c2, c3, c4, c5, c6, c7, c8 = [_clean(row[i]) if i < len(row) else None for i in range(8)]

        c3_str = str(c3) if c3 is not None else ''

        # 汇总行 → 跳过
        if _is_skip_row(c3_str):
            continue

        # BOQ清单项：C1为整数
        try:
            seq = int(c1)
            is_boq = True
        except (TypeError, ValueError):
            is_boq = False

        if is_boq:
            current_item_index = len(items)
            items.append({
                'section_name': current_section_name,
                'item_seq': seq,
                'item_code': str(c2) if c2 is not None else None,
                'item_name': c3_str or None,
                'item_description': str(c4) if c4 is not None else None,
                'unit': str(c5) if c5 is not None else None,
                'quantity': _to_float(c6),
                'unit_price': _to_float(c7),
                'total_price': _to_float(c8),
            })
            continue

        # C1空，C2非空 → 定额子目（含公式或中文特殊子目）
        if c2 is not None and current_item_index is not None:
            c2_str = str(c2).strip()
            quotas.append({
                'item_index': current_item_index,
                'quota_code': c2_str,
                'quota_name': c3_str or None,
                'quota_unit': str(c5) if c5 is not None else None,
                'quantity': _to_float(c6),
                'unit_price': _to_float(c7),
                'total_price': _to_float(c8),
            })
            continue

        # C1空，C2空，C3有名称 → 分部标题
        if c2 is None and c3_str:
            current_section_seq += 1
            current_section_name = c3_str
            sections.append({'seq': current_section_seq, 'section_name': c3_str})

    return {
        'project_name': project_name,
        'bid_section': bid_section,
        'sections': sections,
        'items': items,
        'quotas': quotas,
    }


def _lookup_quota_item_id(conn, code: str) -> int | None:
    if not _SIMPLE_CODE_RE.match(code):
        return None
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM quota_items WHERE item_code = %s LIMIT 1", (code,))
        row = cur.fetchone()
    return row[0] if row else None


def import_to_db(conn, data: dict, source_file: str, tag: str | None, force: bool):
    project_name = data['project_name']

    # 如已存在同文件名工程则删除（force）或跳过
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM manual_boq_projects WHERE source_file = %s", (source_file,))
        existing = cur.fetchone()
    if existing:
        if not force:
            print(f'[跳过] 工程 "{project_name}" 已存在（id={existing[0]}）。使用 --force 强制重新导入。')
            return
        with conn.cursor() as cur:
            cur.execute("DELETE FROM manual_boq_projects WHERE id = %s", (existing[0],))
        conn.commit()
        print(f'[覆盖] 已删除旧记录 id={existing[0]}')

    # 插入工程
    item_count = len(data['items'])
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO manual_boq_projects
                (project_name, bid_section, source_file, tag, item_count)
            VALUES (%s, %s, %s, %s, %s) RETURNING id
        """, (project_name, data['bid_section'], source_file, tag, item_count))
        project_id = cur.fetchone()[0]
    conn.commit()
    print(f'  工程 id={project_id}，{item_count} 条清单项')

    # 插入分部
    section_id_map = {}  # section_name → id
    with conn.cursor() as cur:
        for sec in data['sections']:
            cur.execute("""
                INSERT INTO manual_boq_sections (project_id, seq, section_name)
                VALUES (%s, %s, %s) RETURNING id
            """, (project_id, sec['seq'], sec['section_name']))
            section_id_map[sec['section_name']] = cur.fetchone()[0]
    conn.commit()
    print(f'  {len(section_id_map)} 个分部')

    # 插入清单项
    item_id_list = []
    with conn.cursor() as cur:
        for item in data['items']:
            sec_id = section_id_map.get(item['section_name']) if item['section_name'] else None
            cur.execute("""
                INSERT INTO manual_boq_items
                    (project_id, section_id, item_seq, item_code, item_name,
                     item_description, unit, quantity, unit_price, total_price)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (
                project_id, sec_id, item['item_seq'], item['item_code'],
                item['item_name'], item['item_description'], item['unit'],
                item['quantity'], item['unit_price'], item['total_price'],
            ))
            item_id_list.append(cur.fetchone()[0])
    conn.commit()

    # 插入定额子目
    n_quotas = 0
    n_linked = 0
    with conn.cursor() as cur:
        for q in data['quotas']:
            idx = q['item_index']
            if idx >= len(item_id_list):
                continue
            boq_item_id = item_id_list[idx]
            boq_qty = data['items'][idx]['quantity']

            qty_factor = None
            if q['quantity'] is not None and boq_qty and boq_qty != 0:
                qty_factor = q['quantity'] / boq_qty

            quota_item_id = _lookup_quota_item_id(conn, q['quota_code'])
            if quota_item_id:
                n_linked += 1

            cur.execute("""
                INSERT INTO manual_boq_quotas
                    (boq_item_id, quota_code, quota_name, quota_unit,
                     quantity, unit_price, total_price, qty_factor, quota_item_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                boq_item_id, q['quota_code'], q['quota_name'], q['quota_unit'],
                q['quantity'], q['unit_price'], q['total_price'],
                qty_factor, quota_item_id,
            ))
            n_quotas += 1
    conn.commit()
    print(f'  {n_quotas} 条定额子目，其中 {n_linked} 条成功链接到定额库')


def main():
    parser = argparse.ArgumentParser(description='导入人工套定额工程量清单')
    parser.add_argument('excel_path', help='Excel 文件路径')
    parser.add_argument('--original-name', default=None, help='原始文件名（API上传时使用）')
    parser.add_argument('--tag', default=None, help='工程标签')
    parser.add_argument('--force', action='store_true', help='若已存在则强制覆盖')
    args = parser.parse_args()

    if not os.path.isfile(args.excel_path):
        print(f'[错误] 文件不存在: {args.excel_path}', file=sys.stderr)
        sys.exit(1)

    print(f'解析 {args.excel_path} ...')
    data = parse_workbook(args.excel_path)
    print(f'  工程名：{data["project_name"]}')
    print(f'  {len(data["sections"])} 个分部，{len(data["items"])} 条清单项，{len(data["quotas"])} 条定额子目')

    source_file = args.original_name or os.path.basename(args.excel_path)
    conn = get_connection()
    try:
        print('写入数据库 ...')
        import_to_db(conn, data, source_file, args.tag, args.force)
        print('导入完成。')
    except Exception as e:
        conn.rollback()
        print(f'[错误] {e}', file=sys.stderr)
        raise
    finally:
        conn.close()


if __name__ == '__main__':
    main()
